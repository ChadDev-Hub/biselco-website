from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PrintPageSetup
from fastapi import HTTPException, status
from typing import Optional
import io
def create_technical_report(
    columns:list,
    rows:list,
    prepare_name:Optional[str],
    prepare_position:Optional[str],
    check_name:Optional[str],
    check_position:Optional[str],
    approve_name:Optional[str],
    approve_position:Optional[str],
    title: Optional[str]
):
    # -----------------------------
    # Prepare Data
    # --------------------------------
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="No data available for report")
    ws.title = title if title else "No Title"

     # -----------------------------
    # Styles
    # -----------------------------
    title_font = Font(size=14, bold=True)
    header_font = Font(bold=True, size=12)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    data_font = Font(size=12)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # -----------------------------
    # Title
    # -----------------------------
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(columns))
    ws["A1"] = "CHANGE METER REPORT"
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align

    # -----------------------------
    # Header Row
    # -----------------------------
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # -----------------------------
    # Data Rows
    # -----------------------------
    for row_idx, row in enumerate(rows, start=4):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            center_align_col = [i for i in range(4, len(columns))]
            cell.alignment = center_align if col_idx in center_align_col else left_align
            cell.border = thin_border
            cell.font = data_font   

    # PREPARED BY
    ws.merge_cells(start_row=len(rows) + 2 + 4, start_column=1,
                   end_row=len(rows) + 2 + 4 , end_column=2)
    prepared_by = ws.cell(row=len(rows) + 4 + 2, column=1, value="Prepared by:")
    prepared_by.alignment = left_align
    
    # PREPARED NAME
    ws.merge_cells(start_row=len(rows) + 2 + 4 +2, start_column=1,
                   end_row=len(rows) + 2 + 4 + 2, end_column=2)
    prepared_name = ws.cell(row=len(rows) + 4 + 2 + 2, column=1, value=prepare_name)
    prepared_name.alignment = center_align
    prepared_name.font = Font(size=12, bold=True, underline="single")
    
    # PREPARED POSITION
    ws.merge_cells(start_row=len(rows) + 2 + 4 + 3, start_column=1,
                   end_row=len(rows) + 2 + 4 + 3, end_column=2)
    prepared_position = ws.cell(row=len(rows) + 4 + 2 + 3, column=1, value=prepare_position)
    prepared_position.alignment = center_align
    
    
    # CHECKED BY
    ws.merge_cells(start_row=len(rows) + 2 + 4, start_column=6,
                   end_row=len(rows) + 2 + 4, end_column=7)
    ws.cell(row=len(rows) + 4 + 2, column=6, value="Checked by:")
    
    # CHECKED BY NAME
    ws.merge_cells(start_row=len(rows) + 2 + 4 + 2, start_column=6,
                   end_row=len(rows) + 2 + 4 + 2, end_column=7)
    checked_name = ws.cell(row=len(rows) + 4 + 2 + 2, column=6, value=check_name)
    checked_name.alignment = center_align
    checked_name.font = Font(size=12, bold=True, underline="single")
    
    # CHECKED BY POSITION
    ws.merge_cells(start_row=len(rows) + 2 + 4 + 3, start_column=6,
                   end_row=len(rows) + 2 + 4 + 3, end_column=7)
    checked_position = ws.cell(row=len(rows) + 4 + 2 + 3, column=6, value=check_position)
    checked_position.alignment = center_align
    
    # Approved BY
    ws.merge_cells(start_row=len(rows) + 2 + 4, start_column=10,
                   end_row=len(rows) + 2 + 4, end_column=11)
    ws.cell(row=len(rows) + 4 + 2, column=10, value="Approved by:")
    
    # APPROVED BY NAME
    ws.merge_cells(start_row=len(rows) + 2 + 4 + 2, start_column=10,
                   end_row=len(rows) + 2 + 4 + 2, end_column=11)
    approved_name  = ws.cell(row=len(rows) + 4 + 2 + 2, column=10, value=approve_name)
    approved_name.alignment = center_align
    approved_name.font = Font(size=12, bold=True, underline="single")
    
    # APPROVED BY POSITION
    ws.merge_cells(start_row=len(rows) + 2 + 4 + 3, start_column=10,
                   end_row=len(rows) + 2 + 4 + 3, end_column=11)
    approved_position = ws.cell(row=len(rows) + 4 + 2 + 3, column=10, value=approve_position)
    approved_position.alignment = center_align
    
    # -----------------------------
    # Auto Column Width
    # -----------------------------
    for col_idx, col in enumerate(columns, start=1):
        max_length = len(col)
        for row in rows:
            try:
                max_length = max(max_length, len(str(row[col_idx - 1])))
            except:
                pass
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
    # -----------------------------
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = False

    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    # Print area
    ws.print_title_rows = "1:3"  # repeat header on every page
    # -----------------------------
    # Save
    # -----------------------------
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return file_stream