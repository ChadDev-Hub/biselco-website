from sqlalchemy import select, or_, cast, Text, func, delete, join, true
from sqlalchemy.ext.asyncio.session import AsyncSession
from ..model.change_meter import ChangeMeter
from geoalchemy2.functions import ST_AsGeoJSON
from typing import Optional
from fastapi import HTTPException, status
from openpyxl import Workbook
from ..schema.response_model import ChangeMeterReportResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PrintPageSetup
import io
import json
PAGE_SIZE = 8


async def get_change_meter_stats(session: AsyncSession):
    total_count = (select(
        func.coalesce(func.count(), 0).label("total"))
        .select_from(ChangeMeter)).cte("total_count")
    daily_total = (select(
        func.count().label("daily_total"))
        .select_from(ChangeMeter)
        .where(ChangeMeter.date_accomplished == func.current_date()
               )
    ).cte("daily_total")
    monthly_count = (select(
        func.date_trunc("month", ChangeMeter.date_accomplished).label("month"),
        func.coalesce(func.count(), 0).label("monthly_count"))
        .group_by("month")
    ).subquery()
    average_count = (select(
        func.coalesce(func.floor(func.abs(func.round(
            func.avg(monthly_count.c.monthly_count)))), 0).label("average_count")
    )).cte("average_count")

    data = (await session.execute(
        select(total_count.c.total,
               daily_total.c.daily_total,
               average_count.c.average_count)
        .select_from(total_count)
        .join(daily_total, true())
        .join(average_count, true()))).mappings().one()
    new_data = []
    for key, value in data.items():
        if key == "total":
            total = {
                "title": "Total",
                "value": value,
                "description": "All"
            }
            new_data.append(total)
        elif key == "daily_total":
            daily = {
                "title": "Daily",
                "value": value,
                "description": "Today"
            }
            new_data.append(daily)
        elif key == "average_count":
            average = {
                "title": "Avg.",
                "value": value,
                "description": "Per month avg."
            }
            new_data.append(average)
    return new_data

# GET CHANGE METER DATA BY QUERY OR ALL


async def get_change_meter(session: AsyncSession, query: Optional[str] = None, page: Optional[int] = None):
    if not page:
        page = 1
    change_meter_cte = (select(
        ChangeMeter.id,
        ChangeMeter.timestamped,
        ChangeMeter.date_accomplished,
        ChangeMeter.account_no,
        ChangeMeter.consumer_name,
        ChangeMeter.location,
        ChangeMeter.pull_out_meter,
        ChangeMeter.pull_out_meter_reading,
        ChangeMeter.new_meter_serial_no,
        ChangeMeter.new_meter_brand,
        ChangeMeter.meter_sealed,
        ChangeMeter.initial_reading,
        ChangeMeter.remarks,
        ChangeMeter.accomplished_by,
        func.ceil(func.row_number().over(
            order_by=ChangeMeter.timestamped.desc()) / PAGE_SIZE).label("page"),
        ST_AsGeoJSON(ChangeMeter.geom).label("geom")
    )).cte("change_meter_cte")
    stmt_total = (await session.execute(select(func.count(ChangeMeter.id)))).scalar()
    total_page = 1
    if stmt_total:
        total_page = stmt_total//PAGE_SIZE if stmt_total % PAGE_SIZE == 0 else stmt_total//PAGE_SIZE + 1
    if query:
        stmt = select(change_meter_cte).where(
            or_(
                cast(ChangeMeter.timestamped, Text).ilike(f"%{query}%"),
                cast(ChangeMeter.date_accomplished, Text).ilike(f"%{query}%"),
                ChangeMeter.account_no.ilike(f"%{query}%"),
                ChangeMeter.consumer_name.ilike(f"%{query}%"),
                ChangeMeter.location.ilike(f"%{query}%"),
                ChangeMeter.pull_out_meter.ilike(f"%{query}%"),
                cast(ChangeMeter.pull_out_meter_reading,
                     Text).ilike(f"%{query}%"),
                ChangeMeter.new_meter_serial_no.ilike(f"%{query}%"),
                ChangeMeter.new_meter_brand.ilike(f"%{query}%"),
                cast(ChangeMeter.meter_sealed, Text).ilike(f"%{query}%"),
                cast(ChangeMeter.initial_reading, Text).ilike(f"%{query}%"),
                ChangeMeter.remarks.ilike(f"%{query}%"),
                ChangeMeter.accomplished_by.ilike(f"%{query}%"),
            )).order_by(change_meter_cte.c.timestamped.desc())
    else:
        stmt = select(change_meter_cte).where(change_meter_cte.c.page == page).order_by(
            change_meter_cte.c.timestamped.desc())
    result = (await session.execute(stmt)).mappings().all()
    data = []
    for row in result:
        item = dict(row)
        if item['geom']:
            item['geom'] = json.loads(row['geom'])
        data.append(item)

    change_meter_stats = await get_change_meter_stats(session=session)
    return {
        "data": data,
        "total_page": total_page,
        "stats": change_meter_stats
    }


async def deleteChangeMeter(session: AsyncSession, items: set, page: Optional[int] = None):
    try:
        await session.execute(delete(ChangeMeter).where(ChangeMeter.id.in_(items)))
        await session.commit()
        return await get_change_meter(session=session, page=page)
    except Exception as e:
        await session.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail=str(e))
    finally:
        await session.close()


async def changeMeterReport(session:AsyncSession, 
                            items: set, 
                            approve_name: Optional[str] = None, 
                            approve_position: Optional[str] = None,
                            check_name: Optional[str] = None, 
                            check_position: Optional[str] = None,
                            prepare_name: Optional[str] = None, 
                            prepare_position: Optional[str] = None
                            ):
    stmt = (await session.execute(
        select(ChangeMeter).where(ChangeMeter.id.in_(items))
    )).scalars().all()

    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise HTTPException(status_code=404, detail="No data available for report")
    ws.title = "Change Meter Report"

    # -----------------------------
    # Prepare Data
    # -----------------------------
    data = [
        ChangeMeterReportResponse.model_validate(d).model_dump(mode="json")
        for d in stmt
    ]

    if not data:
        return None

    columns = [col.replace("_", " ").upper() for col in data[0].keys()]
    rows = [list(d.values()) for d in data]

    # -----------------------------
    # Styles
    # -----------------------------
    title_font = Font(size=14, bold=True)
    header_font = Font(bold=True, size=12)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    data_font = Font(size=12)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # -----------------------------
    # Title
    # -----------------------------
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(columns))
    ws["A1"] = "CHANGE METER REPORT"
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align

    # -----------------------------
    # Header Row
    # -----------------------------
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # -----------------------------
    # Data Rows
    # -----------------------------
    for row_idx, row in enumerate(rows, start=4):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = left_align
            cell.border = thin_border
            cell.font = data_font   

    # PREPARED BY
    ws.merge_cells(start_row=len(rows) + 2 + 4, start_column=1,
                   end_row=len(rows) + 2 + 4 , end_column=2)
    prepared_by = ws.cell(row=len(rows) + 4 + 2, column=1, value="Prepared by:")
    prepared_by.alignment = left_align
    
    # PREPARED NAME
    ws.merge_cells(start_row=len(rows) + 2 + 4 +2, start_column=1,
                   end_row=len(rows) + 2 + 4 + 2, end_column=2)
    prepared_name = ws.cell(row=len(rows) + 4 + 2 + 2, column=1, value=prepare_name)
    prepared_name.alignment = center_align
    prepared_name.font = Font(size=12, bold=True, underline="single")
    
    # PREPARED POSITION
    ws.merge_cells(start_row=len(rows) + 2 + 4 + 3, start_column=1,
                   end_row=len(rows) + 2 + 4 + 3, end_column=2)
    prepared_position = ws.cell(row=len(rows) + 4 + 2 + 3, column=1, value=prepare_position)
    prepared_position.alignment = center_align
    
    
    # CHECKED BY
    ws.merge_cells(start_row=len(rows) + 2 + 4, start_column=6,
                   end_row=len(rows) + 2 + 4, end_column=7)
    ws.cell(row=len(rows) + 4 + 2, column=6, value="Checked by:")
    
    # CHECKED BY NAME
    ws.merge_cells(start_row=len(rows) + 2 + 4 + 2, start_column=6,
                   end_row=len(rows) + 2 + 4 + 2, end_column=7)
    checked_name = ws.cell(row=len(rows) + 4 + 2 + 2, column=6, value=check_name)
    checked_name.alignment = center_align
    checked_name.font = Font(size=12, bold=True, underline="single")
    
    # CHECKED BY POSITION
    ws.merge_cells(start_row=len(rows) + 2 + 4 + 3, start_column=6,
                   end_row=len(rows) + 2 + 4 + 3, end_column=7)
    checked_position = ws.cell(row=len(rows) + 4 + 2 + 3, column=6, value=check_position)
    checked_position.alignment = center_align
    
    # Approved BY
    ws.merge_cells(start_row=len(rows) + 2 + 4, start_column=10,
                   end_row=len(rows) + 2 + 4, end_column=11)
    ws.cell(row=len(rows) + 4 + 2, column=10, value="Approved by:")
    
    # APPROVED BY NAME
    ws.merge_cells(start_row=len(rows) + 2 + 4 + 2, start_column=10,
                   end_row=len(rows) + 2 + 4 + 2, end_column=11)
    approved_name  = ws.cell(row=len(rows) + 4 + 2 + 2, column=10, value=approve_name)
    approved_name.alignment = center_align
    approved_name.font = Font(size=12, bold=True, underline="single")
    
    # APPROVED BY POSITION
    ws.merge_cells(start_row=len(rows) + 2 + 4 + 3, start_column=10,
                   end_row=len(rows) + 2 + 4 + 3, end_column=11)
    approved_position = ws.cell(row=len(rows) + 4 + 2 + 3, column=10, value=approve_position)
    approved_position.alignment = center_align
    
    # -----------------------------
    # Auto Column Width
    # -----------------------------
    for col_idx, col in enumerate(columns, start=1):
        max_length = len(col)
        for row in rows:
            try:
                max_length = max(max_length, len(str(row[col_idx - 1])))
            except:
                pass
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
    # -----------------------------
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = False

    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    # Print area
    ws.print_title_rows = "1:3"  # repeat header on every page
    # -----------------------------
    # Save
    # -----------------------------
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return file_stream
